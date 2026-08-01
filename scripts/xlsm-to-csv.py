#!/usr/bin/env python3
"""
Convert a Manuza recipe workbook into the CSV the app's recipe import reads.

Why a conversion step rather than a parser in the browser: the import path is
already built, previewed and tested — matching ingredients, reporting what it
cannot resolve, refusing to invent products. Teaching the browser to read xlsm
would mean a second way in with none of that, and the second one always ends up
the one with the bug. This turns a workbook into the format that path accepts,
so a workbook and a hand-written sheet are the same import.

The layout, one sheet per recipe:

    B1   the word "Category" (a label; the value beside it is often blank)
    B2   dish name
    B4:B29  ingredient rows — name, nett qty, unit, ref %, gross qty, unit,
            cost/unit, cost
    B30  "TOTAL COST"
    C31  yield qty, D31 unit
    G35  price excluding VAT

Sheets that are not recipes — Set Up, Calories, Index, and the blank RECIPE
template — are skipped by name, and any sheet with no dish name in B2 is
skipped for the same reason: a template imported as a dish is a dish full of
empty rows.

    python3 scripts/xlsm-to-csv.py "1 - Recipes.xlsm" recipes.csv
"""
import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover - a setup problem, not a data one
    sys.exit("openpyxl is required: pip3 install openpyxl")

# Sheets that hold configuration or a blank template rather than a dish.
SKIP = {"set up", "calories", "index", "recipe", "sub recipe", "cover"}

FIRST_LINE_ROW = 4
LAST_LINE_ROW = 29


def cell(ws, row, col):
    """Trimmed text, with the workbook's blank-looking "  " treated as blank."""
    v = ws.cell(row, col).value
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in {"", "-"} else s


def number(raw):
    """A figure, or None. Blank and text stay None rather than becoming zero."""
    if raw == "":
        return None
    try:
        return float(str(raw).replace(",", "."))
    except ValueError:
        return None


def extract(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = []
    skipped = []

    for name in wb.sheetnames:
        if name.strip().lower() in SKIP:
            continue
        ws = wb[name]

        dish = cell(ws, 2, 2) or name.strip()
        if not dish:
            skipped.append(name)
            continue

        # The value beside the "Category" label, wherever the sheet puts it.
        category = ""
        for col in (3, 4, 5):
            if cell(ws, 1, col):
                category = cell(ws, 1, col)
                break

        price = number(cell(ws, 35, 7)) or ""

        lines = 0
        for r in range(FIRST_LINE_ROW, LAST_LINE_ROW + 1):
            ingredient = cell(ws, r, 2)
            qty = number(cell(ws, r, 3))
            # A row with no name or no quantity is one of the template's blanks.
            if not ingredient or qty is None:
                continue
            rows.append(
                {
                    "Recipe": dish,
                    "Category": category,
                    "Menu price": price,
                    "Ingredient": ingredient,
                    "Quantity": qty,
                    "Unit": cell(ws, r, 4),
                    "Ref %": number(cell(ws, r, 5)) or 0,
                }
            )
            lines += 1

        if lines == 0:
            skipped.append(name)

    return rows, skipped


def main():
    if len(sys.argv) < 3:
        sys.exit(__doc__)

    source, target = Path(sys.argv[1]), Path(sys.argv[2])
    rows, skipped = extract(source)

    if not rows:
        sys.exit(f"No recipe rows found in {source.name}.")

    fields = ["Recipe", "Category", "Menu price", "Ingredient", "Quantity", "Unit", "Ref %"]
    # utf-8-sig: Excel needs the BOM to read accented ingredient names, and the
    # app's parser strips it.
    with target.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)

    dishes = len({r["Recipe"] for r in rows})
    print(f"{dishes} dishes, {len(rows)} ingredient lines -> {target}")
    if skipped:
        # Named rather than counted: a dish missing from the output because its
        # sheet was laid out differently is worth chasing.
        print(f"skipped {len(skipped)} sheets with no usable rows: "
              + ", ".join(skipped[:10])
              + (f" ... +{len(skipped) - 10}" if len(skipped) > 10 else ""))
    print("Import it from Recipes -> Import recipes. Nothing is created until "
          "you confirm the preview.")


if __name__ == "__main__":
    main()
